import sqlite3
import openpyxl
import os

# Ruta fija para la base de datos
DB_PATH = r"C:\Users\KJhonatan\OneDrive\Escritorio\mi_taller\finanzas.db"

def inicializar_bd():
    # Verifica si la base de datos existe
    if not os.path.exists(DB_PATH):
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute('''
            CREATE TABLE IF NOT EXISTS transacciones (
                id INTEGER PRIMARY KEY,
                fecha TEXT,
                tipo TEXT,
                cantidad REAL,
                descripcion TEXT,
                precio_venta REAL,
                precio_original REAL,
                ganancia REAL,
                monto_total REAL
            )
        ''')
        conn.commit()
        conn.close()
    else:
        print("Base de datos ya existe, no es necesario crearla.")

def agregar_transaccion(fecha, tipo, cantidad, descripcion, precio_venta, precio_original, ganancia, monto_total):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''
        INSERT INTO transacciones (fecha, tipo, cantidad, descripcion, precio_venta, precio_original, ganancia, monto_total)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    ''', (fecha, tipo, cantidad, descripcion, precio_venta, precio_original, ganancia, monto_total))
    conn.commit()
    conn.close()

def obtener_transacciones():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('SELECT * FROM transacciones')
    datos = c.fetchall()
    conn.close()
    return datos

def obtener_transacciones_del_dia():
    from datetime import datetime
    fecha_actual = datetime.now().strftime('%Y-%m-%d')
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('SELECT * FROM transacciones WHERE fecha = ?', (fecha_actual,))
    datos = c.fetchall()
    conn.close()
    return datos

def obtener_datos_producto(identificador):
    """
    Busca un producto en el archivo Excel por índice o por nombre.
    """
    ruta_excel = r"c:\Users\KJhonatan\OneDrive\Escritorio\mi_taller\productos.xlsx"
    wb = openpyxl.load_workbook(ruta_excel)
    hoja = wb.active

    # Leer encabezados de la primera fila
    encabezados = [cell for cell in next(hoja.iter_rows(min_row=1, max_row=1, values_only=True))]

    # Identificar índices de las columnas necesarias
    try:
        idx_indice = encabezados.index("indice")
        idx_nombre = encabezados.index("Nombre")
        idx_precio_venta = encabezados.index("Precio Venta")
        idx_precio_original = encabezados.index("Precio Original")
    except ValueError as e:
        raise ValueError("El archivo Excel no contiene los encabezados esperados.") from e

    # Buscar el producto por índice o nombre
    for fila in hoja.iter_rows(min_row=2, values_only=True):  # Saltar la fila de encabezados
        if str(fila[idx_indice]) == str(identificador) or fila[idx_nombre] == identificador:
            precio_venta = fila[idx_precio_venta]
            precio_original = fila[idx_precio_original]
            return fila[idx_nombre], precio_venta, precio_original

    return None, None, None  # Retorna None si el producto no se encuentra
